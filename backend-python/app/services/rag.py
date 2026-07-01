import os
from typing import List
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from pypdf import PdfReader
import docx
import openpyxl

from app.models import KnowledgeBase
from app.services.openai_service import openai_service

class RAGService:
    def extract_text(self, file_path: str) -> str:
        """
        Extracts text from PDF, DOCX, XLSX, TXT, and MD files.
        """
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".pdf":
            reader = PdfReader(file_path)
            text = ""
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            return text
        elif ext == ".docx":
            doc = docx.Document(file_path)
            return "\n".join([p.text for p in doc.paragraphs])
        elif ext == ".xlsx":
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                lines.append(f"Лист: {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    row_str = " | ".join([str(val) for val in row if val is not None])
                    if row_str.strip():
                        lines.append(row_str)
            return "\n".join(lines)
        elif ext in [".txt", ".md"]:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        else:
            raise ValueError(f"Неподдерживаемый формат файла для RAG: {ext}")

    def chunk_text(self, text: str, chunk_size: int = 800, overlap: int = 150) -> List[str]:
        """
        Splits text into chunks of specified size and overlap.
        """
        chunks: List[str] = []
        if not text:
            return chunks
        
        # Simple character-based splitting
        start = 0
        while start < len(text):
            end = start + chunk_size
            chunk = text[start:end]
            chunks.append(chunk.strip())
            start += chunk_size - overlap
            
        return chunks

    async def ingest_document(self, db: AsyncSession, file_path: str, filename: str):
        """
        Extracts, chunks, embeds and saves a document into the PostgreSQL vector database.
        """
        text = self.extract_text(file_path)
        chunks = self.chunk_text(text)
        
        if not chunks:
            print(f"No text extracted from document {filename}")
            return
            
        for chunk in chunks:
            embedding = await openai_service.get_embedding(chunk)
            kb_entry = KnowledgeBase(
                filename=filename,
                content_chunk=chunk,
                embedding=embedding
            )
            db.add(kb_entry)
            
        await db.commit()
        print(f"Successfully ingested {filename} ({len(chunks)} chunks).")

    async def retrieve_similar_chunks(self, db: AsyncSession, query: str, limit: int = 4) -> List[KnowledgeBase]:
        """
        Retrieves the top N most similar chunks from knowledge base using cosine distance.
        """
        try:
            query_embedding = await openai_service.get_embedding(query)
            
            # Using pgvector's cosine_distance
            distance = KnowledgeBase.embedding.cosine_distance(query_embedding)
            stmt = select(KnowledgeBase).order_by(distance).limit(limit)
            
            result = await db.execute(stmt)
            return list(result.scalars().all())
        except Exception as e:
            print(f"Error retrieving similar chunks: {e}")
            return []

# Instantiate global service
rag_service = RAGService()
